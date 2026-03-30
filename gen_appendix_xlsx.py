# -*- coding: utf-8 -*-
"""Task 4: Generate appendix_table1_dictionary.xlsx"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from dictionaries import PATTERN_TERMS, SYMPTOM_TERMS, ABDOMINAL_TERMS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT = 'analysis_output/revision'

cognition_ja = {}
PTERM_MAP = {'八綱弁証':'pathology','気血津液弁証':'pathology','臓腑弁証':'pathology','六経弁証':'classical'}
for top, subcats in PATTERN_TERMS.items():
    cat = PTERM_MAP.get(top, 'pathology')
    for sub, terms in subcats.items():
        for t in terms:
            cognition_ja[t] = cat
for t in ABDOMINAL_TERMS:
    cognition_ja[t] = 'examination'
extra_ja = {
    '随証':'sho_core','弁証':'sho_core','方証相対':'sho_core',
    '証に基づ':'sho_core','証の変化':'sho_core','証を決定':'sho_core',
    '気血水':'pathology','お血':'pathology','血瘀':'pathology',
    '冷え症':'pathology','冷え性':'pathology','冷え':'pathology',
    'のぼせ':'pathology','気鬱':'pathology','気うつ':'pathology',
    '未病':'epistemological','養生':'epistemological','心身一如':'epistemological',
    '同病異治':'epistemological','異病同治':'epistemological','君臣佐使':'epistemological',
    '傷寒論':'classical','金匱要略':'classical','温病':'classical',
    '転方':'sho_core','奔豚':'pathology','煩躁':'pathology',
    '裏寒':'pathology','表熱':'pathology','血の道':'pathology','水毒':'pathology',
}
cognition_ja.update(extra_ja)
cognition_en = {
    'sho ':'sho_core','sho-based':'sho_core',
    'sho pattern':'sho_core','pattern diagnosis':'sho_core',
    'pattern identification':'sho_core','pattern differentiation':'sho_core',
    'ho-sho-sotai':'sho_core','hoshotai':'sho_core',
    'qi deficiency':'pathology','blood deficiency':'pathology',
    'qi stagnation':'pathology','qi counterflow':'pathology',
    'blood stasis':'pathology','blood stagnation':'pathology',
    'oketsu':'pathology','water toxin':'pathology',
    'fluid disturbance':'pathology','tan-in':'pathology',
    'phlegm-dampness':'pathology','yin deficiency':'pathology',
    'yang deficiency':'pathology','yin-yang':'pathology',
    'ki-ketsu-sui':'pathology','cold sensitivity':'pathology',
    'hie ':'pathology','deficiency pattern':'pathology',
    'excess pattern':'pathology','kyo-jitsu':'pathology',
    'kyojitsu':'pathology','spleen deficiency':'pathology',
    'kidney deficiency':'pathology','liver qi':'pathology',
    'taiyang':'classical','shaoyang':'classical','yangming':'classical',
    'taiyin':'classical','shaoyin':'classical','jueyin':'classical',
    'six stages':'classical','shanghan':'classical','shang han':'classical',
    'jingui':'classical','jin gui yao lue':'classical',
    'abdominal diagnosis':'examination','fukushin':'examination',
    'pulse diagnosis':'examination','tongue diagnosis':'examination',
    'hypochondriac fullness':'examination','kyokyo-kuman':'examination',
    'splashing sound':'examination','abdominal palpation':'examination',
    'mibyou':'epistemological','mibyo':'epistemological',
    'yangsheng':'epistemological','yojo ':'epistemological',
    'mind-body unity':'epistemological',
    'same disease different treatment':'epistemological',
}

liberal_only = set()
for cat, terms in SYMPTOM_TERMS.items():
    if isinstance(terms, list):
        liberal_only.update(terms)
    elif isinstance(terms, dict):
        for subterms in terms.values():
            if isinstance(subterms, list):
                liberal_only.update(subterms)

cat_labels = {
    'sho_core':'T1: Sho Framework',
    'pathology':'T2: Qi-Blood-Water',
    'classical':'T3: Classical Texts',
    'examination':'T4: Kampo Examination',
    'epistemological':'T5: Epistemological',
}

wb = Workbook()
ws = wb.active
ws.title = "Dictionary B (Conservative)"

hf = Font(bold=True, size=11, color='FFFFFF')
hfill = PatternFill('solid', fgColor='4472C4')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

for col, h in enumerate(['Category','Language','Term'], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf; c.fill = hfill; c.border = tb; c.alignment = Alignment(horizontal='center')

row = 2
for cat in ['sho_core','pathology','classical','examination','epistemological']:
    for t in sorted(t for t,c in cognition_ja.items() if c == cat):
        ws.cell(row=row, column=1, value=cat_labels[cat]).border = tb
        ws.cell(row=row, column=2, value='JA').border = tb
        ws.cell(row=row, column=3, value=t).border = tb
        row += 1
    for t in sorted(t for t,c in cognition_en.items() if c == cat):
        ws.cell(row=row, column=1, value=cat_labels[cat]).border = tb
        ws.cell(row=row, column=2, value='EN').border = tb
        ws.cell(row=row, column=3, value=t).border = tb
        row += 1

ws2 = wb.create_sheet("Liberal-only terms")
ws2.cell(row=1, column=1, value='Term').font = Font(bold=True)
ws2.cell(row=1, column=2, value='Type').font = Font(bold=True)
for i, t in enumerate(sorted(liberal_only), 2):
    ws2.cell(row=i, column=1, value=t)
    ws2.cell(row=i, column=2, value='symptom (Liberal only)')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 35

out = os.path.join(OUTPUT, 'appendix_table1_dictionary.xlsx')
wb.save(out)

ja_n = len(cognition_ja); en_n = len(cognition_en)
print(f"Generated: {out}")
print(f"Conservative: JA={ja_n} + EN={en_n} = {ja_n+en_n}")
print(f"Liberal-only: {len(liberal_only)} terms")
for cat in ['sho_core','pathology','classical','examination','epistemological']:
    j = sum(1 for t,c in cognition_ja.items() if c==cat)
    e = sum(1 for t,c in cognition_en.items() if c==cat)
    print(f"  {cat_labels[cat]}: JA={j}, EN={e}")
